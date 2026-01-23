"""
Excel Parser Service
Memory-efficient parser for large Excel files (64MB+)
Supports both .xlsx (openpyxl) and .xls (xlrd via pandas) formats
"""

from datetime import datetime, date
from typing import Generator, Dict, Any, List, Optional
import re
import logging
import pandas as pd

logger = logging.getLogger(__name__)


class ExcelParser:
    """Memory-efficient Excel parser for large files"""
    
    # Header patterns for AUTO-DETECTION (case-insensitive)
    # Each key maps to list of possible header names (will use first match)
    HEADER_PATTERNS = {
        'date': ['дата', 'date'],
        'store_code': ['бсо', 'код точки', 'store_code'],
        'region': ['регион', 'region'],
        'channel': ['канал сбыта', 'канал', 'channel'],
        'store_name': ['адрес', 'address', 'город/точка', 'город'],
        'customer': ['контрагент', 'головной контрагент', 'customer', 'клиент'],
        'category': ['группа', 'группа товара', 'category'],
        'product_type': ['вид товара', 'product_type'],
        'product': ['номенклатура', 'товар', 'product', 'наименование'],
        'barcode': ['штрихкод', 'barcode', 'штрих-код'],
        'quantity': ['количество', 'кол-во', 'qty', 'quantity'],
        'amount': ['сумма с ндс', 'сумма', 'amount', 'total', 'итого'],
    }
    
    # Fallback indices (used only if auto-detection fails)
    FALLBACK_MAP = {
        'date': 3, 'store_code': 4, 'region': 5, 'channel': 6,
        'store_name': 9, 'customer': 8, 'category': 11, 'product_type': 12,
        'product': 13, 'barcode': 14, 'quantity': 16, 'amount': 17,
    }
    
    def __init__(self, file_path: str, chunk_size: int = 5000):
        self.file_path = file_path
        self.chunk_size = chunk_size
        self.total_rows = 0
        self.processed_rows = 0
        self.failed_rows = 0
        self.errors: List[str] = []
        self._use_pandas = False
        self._column_map = {}  # Will be populated by auto-detection
    
    def count_rows(self) -> int:
        """Count total rows in Excel file"""
        try:
            # Try openpyxl first
            from openpyxl import load_workbook
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            ws = wb.active
            self.total_rows = ws.max_row - 1  # Exclude header
            wb.close()
            self._use_pandas = False
            return self.total_rows
        except Exception as e:
            logger.warning(f"openpyxl failed for counting, trying pandas: {e}")
            # Fallback to pandas with multiple engines
            for engine in ['calamine', 'xlrd', 'openpyxl', None]:
                try:
                    logger.info(f"Counting rows with engine: {engine or 'default'}")
                    df = pd.read_excel(self.file_path, header=0, engine=engine)
                    self.total_rows = len(df)
                    self._use_pandas = True
                    logger.info(f"Counted {self.total_rows} rows with {engine or 'default'}")
                    return self.total_rows
                except Exception as e2:
                    logger.warning(f"Engine {engine} failed for counting: {e2}")
                    continue
            logger.error("All engines failed to count rows")
            return 0
    
    def parse_chunks(self) -> Generator[List[Dict[str, Any]], None, None]:
        """Parse Excel file in chunks to save memory"""
        
        if self._use_pandas or self._should_use_pandas():
            yield from self._parse_with_pandas()
        else:
            yield from self._parse_with_openpyxl()
    
    def _detect_columns(self, headers: List[str]) -> None:
        """Auto-detect column indices from headers"""
        self._column_map = {}
        
        # Normalize headers to lowercase for matching
        normalized_headers = [str(h).lower().strip() if h else '' for h in headers]
        
        logger.info(f"Detecting columns from headers: {normalized_headers[:20]}")
        
        for field, patterns in self.HEADER_PATTERNS.items():
            found = False
            for pattern in patterns:
                pattern_lower = pattern.lower()
                for idx, header in enumerate(normalized_headers):
                    if pattern_lower in header or header in pattern_lower:
                        self._column_map[field] = idx
                        logger.info(f"  {field}: matched '{pattern}' at column {idx} ('{headers[idx]}')")
                        found = True
                        break
                if found:
                    break
            
            # Use fallback if not found
            if not found and field in self.FALLBACK_MAP:
                fallback_idx = self.FALLBACK_MAP[field]
                if fallback_idx < len(headers):
                    self._column_map[field] = fallback_idx
                    logger.warning(f"  {field}: using FALLBACK index {fallback_idx}")
        
        # Verify critical columns
        required = ['customer', 'product', 'amount']
        missing = [f for f in required if f not in self._column_map]
        if missing:
            logger.error(f"CRITICAL: Missing required columns: {missing}")
            raise ValueError(f"Excel file missing required columns: {missing}. Headers found: {headers[:15]}")
    
    def _should_use_pandas(self) -> bool:
        """Check if we should use pandas instead of openpyxl"""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            wb.close()
            return False
        except Exception as e:
            logger.warning(f"openpyxl not compatible, using pandas: {e}")
            return True
    
    def _parse_with_pandas(self) -> Generator[List[Dict[str, Any]], None, None]:
        """Parse using pandas for maximum compatibility - tries multiple engines"""
        logger.info(f"Using pandas parser for Excel file: {self.file_path}")
        
        # First try calamine directly (handles corrupted xlsx)
        df = self._try_calamine_direct()
        
        if df is None:
            # Determine file extension
            is_xlsx = str(self.file_path).lower().endswith('.xlsx')
            
            # Try engines in order (openpyxl first for xlsx, xlrd for xls)
            if is_xlsx:
                engines_to_try = ['openpyxl', None]
            else:
                engines_to_try = ['xlrd', 'openpyxl', None]
            
            all_errors = []
            
            for engine in engines_to_try:
                try:
                    logger.info(f"Trying engine: {engine or 'default'}...")
                    df = pd.read_excel(self.file_path, header=0, engine=engine)
                    logger.info(f"SUCCESS with engine: {engine or 'default'}, rows: {len(df)}")
                    break
                except Exception as e:
                    error_msg = f"Engine {engine}: {str(e)[:200]}"
                    all_errors.append(error_msg)
                    logger.error(error_msg)
                    continue
            
            if df is None:
                error_detail = " | ".join(all_errors)
                raise Exception(f"All Excel engines failed: {error_detail}")
        
        self.total_rows = len(df)
        logger.info(f"Loaded {self.total_rows} rows")
        
        # AUTO-DETECT COLUMNS from headers
        self._detect_columns(df.columns.tolist())
        
        # Process in chunks manually
        chunk_num = 0
        parsed_chunk: List[Dict[str, Any]] = []
        
        for idx, row in enumerate(df.itertuples(index=False, name=None)):
            try:
                parsed = self._parse_pandas_row(row, idx + 2)  # +2 for 1-indexed + header
                if parsed:
                    parsed_chunk.append(parsed)
                    self.processed_rows += 1
                else:
                    self.failed_rows += 1
            except Exception as e:
                self.failed_rows += 1
                if len(self.errors) < 100:
                    self.errors.append(f"Row {idx + 2}: {str(e)}")
            
            # Yield chunk when full
            if len(parsed_chunk) >= self.chunk_size:
                chunk_num += 1
                logger.info(f"Chunk {chunk_num}: {self.processed_rows} success, {self.failed_rows} failed")
                yield parsed_chunk
                parsed_chunk = []
            
            # Log progress
            if (idx + 1) % 10000 == 0:
                logger.info(f"Progress: {idx + 1}/{self.total_rows} rows")
        
        # Yield remaining
        if parsed_chunk:
            yield parsed_chunk
        
        logger.info(f"Pandas parsing complete: {self.processed_rows} success, {self.failed_rows} failed")
    
    def _try_calamine_direct(self) -> Optional[pd.DataFrame]:
        """Try reading with calamine directly for corrupted xlsx files"""
        try:
            from python_calamine import CalamineWorkbook
            logger.info("Trying calamine direct read...")
            
            workbook = CalamineWorkbook.from_path(self.file_path)
            sheet_names = workbook.sheet_names
            if not sheet_names:
                logger.warning("Calamine: no sheets found")
                return None
            
            # Read first sheet
            data = workbook.get_sheet_by_name(sheet_names[0]).to_python()
            if not data or len(data) < 2:
                logger.warning("Calamine: no data found")
                return None
            
            # First row is header
            headers = [str(h) if h else f"col_{i}" for i, h in enumerate(data[0])]
            rows = data[1:]
            
            df = pd.DataFrame(rows, columns=headers)
            logger.info(f"Calamine SUCCESS: {len(df)} rows, {len(headers)} columns")
            return df
            
        except ImportError:
            logger.warning("Calamine not available")
            return None
        except Exception as e:
            logger.warning(f"Calamine direct read failed: {e}")
            return None
    
    def _parse_pandas_row(self, row: Any, row_num: int) -> Optional[Dict[str, Any]]:
        """Parse a pandas row (Series or tuple) into structured data"""
        
        # Get values by column index
        if hasattr(row, 'values'):
            cols = row.values
        else:
            cols = row
        
        # Use auto-detected column map with safe fallbacks
        cmap = self._column_map
        
        date_val = self._safe_get(cols, cmap.get('date', -1))
        customer = self._safe_get(cols, cmap.get('customer', -1))
        product = self._safe_get(cols, cmap.get('product', -1))
        quantity = self._safe_get(cols, cmap.get('quantity', -1))
        amount = self._safe_get(cols, cmap.get('amount', -1))
        
        # Validate required fields
        if pd.isna(customer) or pd.isna(product):
            return None
        
        customer = str(customer)
        product = str(product)
        
        if not customer.strip() or not product.strip():
            return None
        
        # Parse date
        sale_date = self._parse_date(date_val)
        if not sale_date:
            return None
        
        # Parse numbers
        qty = self._parse_number(quantity)
        amt = self._parse_number(amount)
        
        if amt is None or amt <= 0:
            return None
        
        # Calculate price per unit
        price = amt / qty if qty and qty > 0 else amt
        
        return {
            'row_num': row_num,
            'sale_date': sale_date,
            'customer_name': self._normalize_name(customer),
            'customer_raw': customer,
            'product_name': self._normalize_name(product),
            'product_raw': product,
            'category': str(self._safe_get(cols, cmap.get('category', -1)) or 'Без категории'),
            'store_code': self._safe_get(cols, cmap.get('store_code', -1)),
            'store_name': self._safe_get(cols, cmap.get('store_name', -1)),
            'region': self._safe_get(cols, cmap.get('region', -1)),
            'channel': self._safe_get(cols, cmap.get('channel', -1)),
            'quantity': qty or 1,
            'price': round(price, 2),
            'amount': round(amt, 2),
            'year': sale_date.year,
            'month': sale_date.month,
            'week': sale_date.isocalendar()[1],
            'day_of_week': sale_date.weekday()
        }
    
    def _safe_get(self, arr, idx: int) -> Any:
        """Safely get value from array"""
        try:
            if idx < len(arr):
                val = arr[idx]
                if pd.isna(val):
                    return None
                return val
            return None
        except:
            return None
    
    def _parse_with_openpyxl(self) -> Generator[List[Dict[str, Any]], None, None]:
        """Parse using openpyxl (original method)"""
        logger.info("Using openpyxl parser for Excel file")
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            ws = wb.active
            
            # Get headers from first row for auto-detection
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(h) if h else f'col_{i}' for i, h in enumerate(header_row)]
            self._detect_columns(headers)
            
            chunk: List[Dict[str, Any]] = []
            row_num = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
                row_num += 1
                
                try:
                    parsed = self._parse_row(row, row_num)
                    if parsed:
                        chunk.append(parsed)
                        self.processed_rows += 1
                    else:
                        self.failed_rows += 1
                except Exception as e:
                    self.failed_rows += 1
                    if len(self.errors) < 100:
                        self.errors.append(f"Row {row_num}: {str(e)}")
                
                # Yield chunk when full
                if len(chunk) >= self.chunk_size:
                    logger.info(f"Parsed {row_num} rows...")
                    yield chunk
                    chunk = []
                
                # Log progress every 10,000 rows
                if row_num % 10000 == 0:
                    logger.info(f"Progress: {row_num} rows processed, {self.failed_rows} failed")
            
            # Yield remaining rows
            if chunk:
                yield chunk
            
            wb.close()
            logger.info(f"Parsing complete: {self.processed_rows} success, {self.failed_rows} failed")
            
        except Exception as e:
            logger.error(f"openpyxl parsing error: {e}")
            raise
    
    def _parse_row(self, row: tuple, row_num: int) -> Optional[Dict[str, Any]]:
        """Parse a single row into structured data"""
        
        # Use auto-detected column map
        cmap = self._column_map
        
        date_val = self._get_cell(row, cmap.get('date', -1))
        customer = self._get_cell(row, cmap.get('customer', -1))
        product = self._get_cell(row, cmap.get('product', -1))
        quantity = self._get_cell(row, cmap.get('quantity', -1))
        amount = self._get_cell(row, cmap.get('amount', -1))
        
        # Validate required fields
        if not customer or not product:
            return None
        
        # Parse date
        sale_date = self._parse_date(date_val)
        if not sale_date:
            return None
        
        # Parse numbers
        qty = self._parse_number(quantity)
        amt = self._parse_number(amount)
        
        if amt is None or amt <= 0:
            return None
        
        # Calculate price per unit
        price = amt / qty if qty and qty > 0 else amt
        
        return {
            'row_num': row_num,
            'sale_date': sale_date,
            'customer_name': self._normalize_name(str(customer)),
            'customer_raw': str(customer),
            'product_name': self._normalize_name(str(product)),
            'product_raw': str(product),
            'category': self._get_cell(row, cmap.get('category', -1)) or 'Без категории',
            'store_code': self._get_cell(row, cmap.get('store_code', -1)),
            'store_name': self._get_cell(row, cmap.get('store_name', -1)),
            'region': self._get_cell(row, cmap.get('region', -1)),
            'channel': self._get_cell(row, cmap.get('channel', -1)),
            'quantity': qty or 1,
            'price': round(price, 2),
            'amount': round(amt, 2),
            'year': sale_date.year,
            'month': sale_date.month,
            'week': sale_date.isocalendar()[1],
            'day_of_week': sale_date.weekday()
        }
    
    def _get_cell(self, row: tuple, col_index: int) -> Any:
        """Safely get cell value by index"""
        try:
            if col_index < len(row):
                val = row[col_index]
                return val if val is not None else None
            return None
        except:
            return None
    
    def _parse_date(self, value: Any) -> Optional[date]:
        """Parse date from various formats"""
        if value is None:
            return None
        
        # Handle pandas Timestamp
        if hasattr(value, 'date'):
            try:
                return value.date()
            except:
                pass
        
        # Already a date/datetime
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        
        # String parsing
        if isinstance(value, str):
            value = value.strip()
            
            # Try various formats
            for fmt in ['%d.%m.%Y', '%d.%m.%y', '%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d']:
                try:
                    return datetime.strptime(value, fmt).date()
                except:
                    continue
        
        # Try parsing as number (Excel serial date)
        if isinstance(value, (int, float)):
            try:
                # Excel date starts from 1899-12-30
                from datetime import timedelta
                return (datetime(1899, 12, 30) + timedelta(days=int(value))).date()
            except:
                pass
        
        return None
    
    def _parse_number(self, value: Any) -> Optional[float]:
        """Parse number from string or number"""
        if value is None:
            return None
        
        if isinstance(value, (int, float)):
            if pd.isna(value):
                return None
            return float(value)
        
        if isinstance(value, str):
            # Remove spaces, replace comma with dot
            value = value.strip().replace(' ', '').replace(',', '.')
            try:
                return float(value)
            except:
                return None
        
        return None
    
    # Compiled patterns for normalization
    _REMOVE_PATTERNS = [
        re.compile(p) for p in [
            r'^ооо\s+',
            r'^оао\s+',
            r'^зао\s+',
            r'^ип\s+',
            r'^чуп\s+',
            r'^уп\s+',
            r'\s+ооо$',
            r'\s+оао$',
            r'"',
            r"'",
        ]
    ]

    def _normalize_name(self, name: str) -> str:
        """Normalize company/product name for deduplication"""
        if not name:
            return ''
        
        # Convert to lowercase
        normalized = name.lower().strip()

        # Remove common prefixes/suffixes
        for pattern in self._REMOVE_PATTERNS:
            normalized = pattern.sub('', normalized)
        
        # Remove extra whitespace
        normalized = ' '.join(normalized.split())
        
        return normalized
    
    def get_stats(self) -> Dict[str, Any]:
        """Get parsing statistics"""
        return {
            'total_rows': self.total_rows,
            'processed_rows': self.processed_rows,
            'failed_rows': self.failed_rows,
            'success_rate': round(self.processed_rows / max(self.total_rows, 1) * 100, 1),
            'errors': self.errors[:20]  # First 20 errors
        }
