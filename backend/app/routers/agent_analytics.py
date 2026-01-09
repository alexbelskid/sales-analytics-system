"""
Agent Analytics API Router
REST API endpoints for agent performance tracking and analytics
"""

from fastapi import APIRouter, Query, HTTPException, UploadFile, File
from datetime import date, datetime, timedelta
from typing import Optional, List
from uuid import UUID
import logging
import io
import csv
import openpyxl

from app.models.agent_analytics import (
    AgentDashboardMetrics,
    AgentPerformance,
    AgentPerformanceDetailed,
    RegionalPerformance,
    AgentSalesPlanCreate,
    AgentDailySalesCreate,
    GoogleSheetsImportResult,
)
from app.services.agent_analytics_service import agent_analytics_service
from app.services.google_sheets_importer import google_sheets_importer

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/agent-analytics", tags=["agent-analytics"])


# ============================================================================
# Dashboard & Overview
# ============================================================================

@router.get("/dashboard", response_model=AgentDashboardMetrics)
async def get_agent_dashboard(
    period_start: Optional[date] = Query(None, description="Start date of period"),
    period_end: Optional[date] = Query(None, description="End date of period"),
    region: Optional[str] = Query(None, description="Filter by region")
):
    """
    Get comprehensive dashboard metrics for agent analytics
    
    Returns overall metrics, regional breakdowns, and top/bottom performers
    """
    try:
        # Default to current month if not specified
        if not period_start or not period_end:
            today = date.today()
            period_start = date(today.year, today.month, 1)
            # Last day of current month
            if today.month == 12:
                period_end = date(today.year, 12, 31)
            else:
                period_end = date(today.year, today.month + 1, 1) - timedelta(days=1)
        
        return await agent_analytics_service.get_dashboard_metrics(
            period_start, period_end, region
        )
    
    except Exception as e:
        logger.error(f"Dashboard error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# Agent Performance
# ============================================================================

@router.get("/agents", response_model=List[AgentPerformance])
async def get_all_agents(
    period_start: Optional[date] = Query(None),
    period_end: Optional[date] = Query(None),
    region: Optional[str] = Query(None, description="Filter by region"),
    min_fulfillment: Optional[float] = Query(None, ge=0, description="Minimum fulfillment %"),
    max_fulfillment: Optional[float] = Query(None, le=200, description="Maximum fulfillment %"),
):
    """Get performance metrics for all agents with optional filters"""
    try:
        # Default to current month
        if not period_start or not period_end:
            today = date.today()
            period_start = date(today.year, today.month, 1)
            if today.month == 12:
                period_end = date(today.year, 12, 31)
            else:
                period_end = date(today.year, today.month + 1, 1) - timedelta(days=1)
        
        return await agent_analytics_service.get_all_agents_performance(
            period_start, period_end, region, min_fulfillment, max_fulfillment
        )
    
    except Exception as e:
        logger.error(f"Error fetching agents: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/agents/{agent_id}", response_model=AgentPerformanceDetailed)
async def get_agent_details(
    agent_id: UUID,
    period_start: Optional[date] = Query(None),
    period_end: Optional[date] = Query(None),
):
    """Get detailed performance metrics for a specific agent"""
    try:
        # Default to current month
        if not period_start or not period_end:
            today = date.today()
            period_start = date(today.year, today.month, 1)
            if today.month == 12:
                period_end = date(today.year, 12, 31)
            else:
                period_end = date(today.year, today.month + 1, 1) - timedelta(days=1)
        
        return await agent_analytics_service.get_agent_performance(
            agent_id, period_start, period_end, detailed=True
        )
    
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.error(f"Error fetching agent {agent_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# Regional Analytics
# ============================================================================

@router.get("/regions", response_model=List[RegionalPerformance])
async def get_all_regions(
    period_start: Optional[date] = Query(None),
    period_end: Optional[date] = Query(None),
):
    """Get performance metrics for all regions"""
    try:
        # Default to current month
        if not period_start or not period_end:
            today = date.today()
            period_start = date(today.year, today.month, 1)
            if today.month == 12:
                period_end = date(today.year, 12, 31)
            else:
                period_end = date(today.year, today.month + 1, 1) - timedelta(days=1)
        
        # Get all unique regions
        regions = ['БРЕСТ', 'ВИТЕБСК', 'ГОМЕЛЬ', 'ГРОДНО', 'МИНСК']
        
        results = []
        for region in regions:
            try:
                perf = await agent_analytics_service.get_regional_performance(
                    region, period_start, period_end
                )
                if perf.agents_count > 0:  # Only include regions with agents
                    results.append(perf)
            except Exception as region_error:
                logger.warning(f"Error fetching region {region}: {region_error}")
                continue
        
        return results
    
    except Exception as e:
        logger.error(f"Error fetching regions: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/regions/{region}", response_model=RegionalPerformance)
async def get_region_details(
    region: str,
    period_start: Optional[date] = Query(None),
    period_end: Optional[date] = Query(None),
):
    """Get detailed performance metrics for a specific region"""
    try:
        # Default to current month
        if not period_start or not period_end:
            today = date.today()
            period_start = date(today.year, today.month, 1)
            if today.month == 12:
                period_end = date(today.year, 12, 31)
            else:
                period_end = date(today.year, today.month + 1, 1) - timedelta(days=1)
        
        return await agent_analytics_service.get_regional_performance(
            region, period_start, period_end
        )
    
    except Exception as e:
        logger.error(f"Error fetching region {region}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# Rankings
# ============================================================================

@router.get("/rankings", response_model=List[AgentPerformance])
async def get_agent_rankings(
    period_start: Optional[date] = Query(None),
    period_end: Optional[date] = Query(None),
    region: Optional[str] = Query(None, description="Filter by region"),
    limit: int = Query(20, ge=1, le=100, description="Number of top agents to return")
):
    """Get agent rankings sorted by performance"""
    try:
        # Default to current month
        if not period_start or not period_end:
            today = date.today()
            period_start = date(today.year, today.month, 1)
            if today.month == 12:
                period_end = date(today.year, 12, 31)
            else:
                period_end = date(today.year, today.month + 1, 1) - timedelta(days=1)
        
        agents = await agent_analytics_service.get_all_agents_performance(
            period_start, period_end, region
        )
        
        # Sort by fulfillment percentage descending
        sorted_agents = sorted(agents, key=lambda x: x.fulfillment_percent, reverse=True)
        
        # Add rankings
        for idx, agent in enumerate(sorted_agents, 1):
            agent.ranking = idx
        
        return sorted_agents[:limit]
    
    except Exception as e:
        logger.error(f"Error fetching rankings: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# Data Import
# ============================================================================

@router.post("/import-excel", response_model=GoogleSheetsImportResult)
async def import_from_excel(
    file: UploadFile = File(..., description="Excel or CSV file with agent sales data"),
    period_start: date = Query(..., description="Start date of period"),
    period_end: date = Query(..., description="End date of period"),
):
    """
    Import agent sales data from Excel (.xlsx) or CSV file (Daily Sales-Out format)
    
    Expected format:
    - Regional headers (БРЕСТ, ВИТЕБСК, etc.)
    - Agent rows: Name | Sales | Plan | Fulfillment% | Forecast% | Categories... | Daily history...
    """
    try:
        contents = await file.read()
        filename = file.filename or ""
        data = []
        
        # Determine file type and parse accordingly
        if filename.lower().endswith('.csv'):
            # Parse CSV file
            try:
                text_content = contents.decode('utf-8')
            except UnicodeDecodeError:
                text_content = contents.decode('windows-1251')  # Cyrillic encoding fallback
            
            reader = csv.reader(io.StringIO(text_content))
            data = list(reader)
            logger.info(f"Parsed CSV with {len(data)} rows")
            
        elif filename.lower().endswith(('.xlsx', '.xls')):
            # Parse Excel file
            try:
                workbook = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
                sheet = workbook.active
                for row in sheet.iter_rows(values_only=True):
                    data.append(list(row))
                logger.info(f"Parsed Excel with {len(data)} rows")
            except Exception as excel_error:
                logger.error(f"Excel parse error: {excel_error}")
                raise HTTPException(
                    status_code=400,
                    detail=f"Ошибка чтения Excel файла. Убедитесь, что файл имеет формат .xlsx. Ошибка: {str(excel_error)}"
                )
        else:
            # Try to auto-detect format
            try:
                # Try Excel first
                workbook = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
                sheet = workbook.active
                for row in sheet.iter_rows(values_only=True):
                    data.append(list(row))
            except:
                # Try CSV
                try:
                    text_content = contents.decode('utf-8')
                except UnicodeDecodeError:
                    text_content = contents.decode('windows-1251')
                reader = csv.reader(io.StringIO(text_content))
                data = list(reader)
        
        if not data:
            raise HTTPException(
                status_code=400,
                detail="Файл пуст или не содержит данных"
            )
        
        logger.info(f"Starting import of {len(data)} rows for period {period_start} to {period_end}")
        
        # Import using the importer service
        result = await google_sheets_importer.import_from_data(
            data, period_start, period_end
        )
        
        return result
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Import error: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка импорта: {str(e)}. Поддерживаемые форматы: .xlsx, .csv"
        )


# ============================================================================
# Plans and Daily Sales Management
# ============================================================================

@router.post("/plans")
async def create_sales_plan(plan: AgentSalesPlanCreate):
    """Create a sales plan for an agent"""
    try:
        from app.database import supabase
        
        result = supabase.table("agent_sales_plans").insert({
            "agent_id": str(plan.agent_id),
            "period_start": plan.period_start.isoformat(),
            "period_end": plan.period_end.isoformat(),
            "plan_amount": plan.plan_amount,
            "region": plan.region
        }).execute()
        
        return {"success": True, "data": result.data[0] if result.data else None}
    
    except Exception as e:
        logger.error(f"Error creating plan: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/daily-sales")
async def create_daily_sales(sales: AgentDailySalesCreate):
    """Record daily sales for an agent"""
    try:
        from app.database import supabase
        
        result = supabase.table("agent_daily_sales").insert({
            "agent_id": str(sales.agent_id),
            "sale_date": sales.sale_date.isoformat(),
            "amount": sales.amount,
            "region": sales.region,
            "category": sales.category
        }).execute()
        
        return {"success": True, "data": result.data[0] if result.data else None}
    
    except Exception as e:
        logger.error(f"Error creating daily sales: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# Utility Endpoints
# ============================================================================

@router.get("/health")
async def health_check():
    """Health check endpoint"""
    return {
        "status": "healthy",
        "service": "agent-analytics",
        "timestamp": datetime.now().isoformat()
    }
